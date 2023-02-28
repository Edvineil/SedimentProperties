# This is main.py
import os
import pandas as pd
import openpyxl as oxl
from openpyxl.styles import Font, Alignment, PatternFill
from water import WaterParameters
from fun import *
from settling import *


# function that inserts together all outputs (sediments and water properties) in a new data frame.
def join_results(diameters, velocity, u_coefficient, u_condition, w_density, w_viscosity):
    output = pd.merge(diameters, velocity, how='outer', left_index=True, right_index=True)
    output = output.reindex(
        columns=output.columns.tolist() + ["U_coeff (-)", "W_density (kg/m3)", 'W_viscosity (m2/s)'])
    output.iat[0, 2] = u_coefficient
    output.iat[1, 2] = u_condition
    output.iat[0, 3] = w_density
    output.iat[0, 4] = w_viscosity
    return output


# function that implements new cell style in the headers of the output sed_properties.xlsx file.
def header_style(xlsx_file_name):
    # implement new cell styles in the headers
    w_book = oxl.load_workbook(filename=xlsx_file_name, read_only=False)
    w_sheet = w_book.active

    # define title styles
    title_font = Font(name="Arial", size="10", bold=True, italic=True, color="C1D0DE")
    title_fill = PatternFill(fill_type="solid", start_color="1B7DE9", end_color="2027DA")
    title_align = Alignment(horizontal='center', vertical='bottom', text_rotation=0,
                            wrap_text=False, shrink_to_fit=False, indent=0)

    # write title styles
    for row in w_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
        for cell in row:
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_align

    w_book.save(os.path.abspath("..") + "\\sed_properties.xlsx")
    return w_book.close()


@log_actions
def main():
    # get water properties with a given temperature in °C
    water = WaterParameters(temperature=20)

    # calculate the characteristic grain sizes
    GrainsReader(csv_file_name=os.path.abspath("..") + "\\grains.csv")
    grains = GradationParameters(show_plot=True)

    # calculate the settling velocity and save the sediment properties as an Excel file
    velocity = SettlingParameters(viscosity=water.v)
    sed_properties = join_results(grains.D_char, velocity.ws, grains.U, grains.condition, water.p, water.v)
    sed_properties.to_excel(os.path.abspath("..") + "\\sed_properties.xlsx")

    # new cell styles in the headers of the sediment properties file
    header_style(xlsx_file_name=os.path.abspath("..") + "\\sed_properties.xlsx")


if __name__ == '__main__':
    main()
